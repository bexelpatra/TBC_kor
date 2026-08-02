"""엑셀 일괄 등록 — 양식 생성 / 파싱."""
import io
import re

from openpyxl import Workbook, load_workbook

HEADERS = ["뒷번호(학부모)", "이름", "번호(4자리)"]


TEMPLATE_ROWS = 1000  # 입력 영역에 텍스트 서식을 미리 적용해둘 행 수


def build_template() -> bytes:
    """다운로드용 엑셀 양식(헤더 + 예시 1행).
    뒷번호·번호 컬럼은 텍스트(@) 서식으로 지정 — 0으로 시작하는 값이
    엑셀에서 숫자로 변환되어 앞자리 0이 사라지는 것을 방지.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "학생목록"
    ws.append(HEADERS)
    ws.append(["1234", "홍길동", "5678"])  # 예시
    for row in ws.iter_rows(min_row=1, max_row=TEMPLATE_ROWS, min_col=1, max_col=1):
        row[0].number_format = "@"
    for row in ws.iter_rows(min_row=1, max_row=TEMPLATE_ROWS, min_col=3, max_col=3):
        row[0].number_format = "@"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_rows(data: bytes) -> list[dict]:
    """업로드 엑셀 → 행 리스트. 검증은 호출부에서.
    반환: [{row, phone_tail, name, student_number}] (row=엑셀 행 번호, 2부터)
    """
    wb = load_workbook(io.BytesIO(data), read_only=True)
    ws = wb.active
    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None for c in row):
            continue
        phone, name, student_number = (list(row) + [None, None, None])[:3]
        rows.append({
            "row": idx,
            "phone_tail": str(phone).strip() if phone is not None else "",
            "name": str(name).strip() if name is not None else "",
            "student_number": str(student_number).strip() if student_number is not None else "",
        })
    return rows


def validate_row(r: dict) -> str | None:
    """유효성 검사. 통과면 None, 실패면 사유 문자열."""
    if not re.fullmatch(r"\d{4}", r["phone_tail"]):
        return "뒷번호는 숫자 4자리"
    if not (1 <= len(r["name"]) <= 20):
        return "이름은 1~20자"
    if not re.fullmatch(r"\d{4}", r["student_number"]):
        return "번호는 숫자 4자리"
    return None


RESULT_HEADERS = ["행", "상태", "사유"]


def build_result_report(results: list[dict]) -> bytes:
    """일괄 등록 결과 전체를 엑셀로. results: [{row, status, reason}]."""
    wb = Workbook()
    ws = wb.active
    ws.title = "처리결과"
    ws.append(RESULT_HEADERS)
    for r in results:
        ws.append([r["row"], r["status"], r.get("reason") or ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
